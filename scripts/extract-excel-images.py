#!/usr/bin/env python3
"""
Extract embedded images from Excel files and output a JSON mapping.

Usage:
    python extract-excel-images.py <input.xls|xlsx> <output_dir>

Output (JSON to stdout):
    {
      "sheetName": {
        "row_col": "/uploads/question-images/<filename>.png",
        ...
      }
    }

If input is .xls, it will be converted to .xlsx using LibreOffice first.
"""

import sys
import os
import json
import subprocess
import tempfile
import hashlib
from pathlib import Path

def find_libreoffice():
    """Find LibreOffice binary."""
    candidates = [
        'libreoffice',
        'soffice',
        r'C:\Program Files\LibreOffice\program\soffice.exe',
        r'C:\Program Files (x86)\LibreOffice\program\soffice.exe',
        '/usr/bin/libreoffice',
        '/usr/bin/soffice',
    ]
    for cmd in candidates:
        if os.path.isfile(cmd):
            return cmd
    # Try which
    for cmd in ['libreoffice', 'soffice']:
        try:
            result = subprocess.run(['which', cmd], capture_output=True, text=True)
            if result.returncode == 0 and result.stdout.strip():
                return result.stdout.strip()
        except Exception:
            pass
    return None

def convert_xls_to_xlsx(xls_path: str) -> str:
    """Convert .xls to .xlsx using LibreOffice. Returns path to .xlsx file."""
    lo = find_libreoffice()
    if not lo:
        raise RuntimeError('LibreOffice not found. Install it to support .xls image extraction.')

    tmp_dir = tempfile.mkdtemp()
    subprocess.run(
        [lo, '--headless', '--convert-to', 'xlsx', '--outdir', tmp_dir, xls_path],
        capture_output=True,
        timeout=30,
    )

    # Find the output file
    base = Path(xls_path).stem
    xlsx_path = os.path.join(tmp_dir, base + '.xlsx')
    if not os.path.isfile(xlsx_path):
        # Try to find any xlsx in the temp dir
        for f in os.listdir(tmp_dir):
            if f.endswith('.xlsx'):
                xlsx_path = os.path.join(tmp_dir, f)
                break
        else:
            raise RuntimeError(f'LibreOffice conversion failed. No .xlsx found in {tmp_dir}')

    return xlsx_path

def extract_images(xlsx_path: str, output_dir: str) -> dict:
    """Extract images from .xlsx file and save to output_dir."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError('openpyxl not installed. Run: pip install openpyxl')

    wb = load_workbook(xlsx_path)
    os.makedirs(output_dir, exist_ok=True)

    result = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        images = ws._images
        if not images:
            continue

        sheet_result = {}
        for img in images:
            anchor = img.anchor
            if not hasattr(anchor, '_from'):
                continue

            fr = anchor._from
            row = fr.row    # 0-indexed
            col = fr.col    # 0-indexed

            # Generate a unique filename based on content hash
            img_data = img._data()
            content_hash = hashlib.md5(img_data).hexdigest()[:8]
            filename = f'{sheet_name}_r{row}_c{col}_{content_hash}.png'
            filepath = os.path.join(output_dir, filename)

            with open(filepath, 'wb') as f:
                f.write(img_data)

            # Key format: "row_col" (0-indexed)
            key = f'{row}_{col}'
            # URL path relative to public/
            url_path = '/uploads/question-images/' + filename
            sheet_result[key] = url_path

        if sheet_result:
            result[sheet_name] = sheet_result

    return result

def main():
    if len(sys.argv) < 3:
        print(json.dumps({'error': 'Usage: extract-excel-images.py <input> <output_dir>'}))
        sys.exit(1)

    input_path = sys.argv[1]
    output_dir = sys.argv[2]

    if not os.path.isfile(input_path):
        print(json.dumps({'error': f'File not found: {input_path}'}))
        sys.exit(1)

    try:
        # Convert .xls to .xlsx if needed
        if input_path.lower().endswith('.xls') and not input_path.lower().endswith('.xlsx'):
            xlsx_path = convert_xls_to_xlsx(input_path)
            cleanup_xlsx = True
        else:
            xlsx_path = input_path
            cleanup_xlsx = False

        # Extract images
        mapping = extract_images(xlsx_path, output_dir)

        # Cleanup temp file
        if cleanup_xlsx:
            try:
                os.unlink(xlsx_path)
                os.rmdir(os.path.dirname(xlsx_path))
            except Exception:
                pass

        print(json.dumps(mapping, ensure_ascii=False))

    except Exception as e:
        print(json.dumps({'error': str(e)}))
        sys.exit(1)

if __name__ == '__main__':
    main()
